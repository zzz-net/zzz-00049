"""文件解析模块 - 支持银行回单、系统流水、手工调整三类 CSV/XLSX."""
from __future__ import annotations

import csv
import os
from dataclasses import dataclass, field
from typing import Dict, List, Tuple, Any, Optional, Iterator

from .models import FileType, Transaction, ImportedFile
from .config import get_column_aliases, load_config


DEFAULT_COLUMN_MAPS = {
    FileType.BANK_STATEMENT: {
        "txn_id": ["交易流水号", "流水号", "transaction_id", "txn_id"],
        "amount": ["金额", "交易金额", "amount"],
        "date": ["交易日期", "日期", "date", "transaction_date"],
        "counterparty": ["对方户名", "对方账户名", "交易对手", "counterparty"],
        "description": ["摘要", "备注", "用途", "description", "summary"],
        "currency": ["币种", "currency"],
    },
    FileType.SYSTEM_RECEIPT: {
        "txn_id": ["订单号", "交易号", "收款流水号", "transaction_id", "txn_id", "order_id"],
        "amount": ["收款金额", "金额", "amount"],
        "date": ["收款日期", "日期", "date", "receipt_date"],
        "counterparty": ["付款方", "客户", "customer", "counterparty"],
        "description": ["备注", "说明", "description"],
        "currency": ["币种", "currency"],
    },
    FileType.MANUAL_ADJUSTMENT: {
        "txn_id": ["调整编号", "流水号", "transaction_id", "txn_id", "adjustment_id"],
        "amount": ["调整金额", "金额", "amount"],
        "date": ["调整日期", "日期", "date"],
        "counterparty": ["对方", "counterparty"],
        "description": ["调整原因", "备注", "description", "reason"],
        "currency": ["币种", "currency"],
    },
}


@dataclass
class ParseError:
    """解析错误记录."""
    source_file: str
    source_row: int
    error_type: str
    message: str
    raw_row: Dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> Dict[str, Any]:
        return {
            "source_file": self.source_file,
            "source_row": self.source_row,
            "error_type": self.error_type,
            "message": self.message,
            "raw_row": self.raw_row,
        }


@dataclass
class ParseResult:
    """解析结果."""
    transactions: List[Transaction] = field(default_factory=list)
    errors: List[ParseError] = field(default_factory=list)
    duplicates: List[ParseError] = field(default_factory=list)

    @property
    def row_count(self) -> int:
        return len(self.transactions)

    @property
    def error_count(self) -> int:
        return len(self.errors) + len(self.duplicates)


def _detect_column(header: List[str], candidates: List[str]) -> Optional[str]:
    """根据候选列名列表自动检测实际列名."""
    header_lower = {h.strip().lower(): h for h in header}
    for cand in candidates:
        if cand in header:
            return cand
        if cand.lower() in header_lower:
            return header_lower[cand.lower()]
    return None


def _build_column_map(
    file_type: FileType,
    header: List[str],
    user_aliases: Optional[Dict[str, str]] = None,
    extra_col_map: Optional[Dict[str, str]] = None,
) -> Dict[str, str]:
    """构建字段名到CSV列名的映射.

    优先级:
    1. 命令行传入的 --col-map 参数 (extra_col_map: {别名列名: 标准字段})
    2. 用户在 config.yaml 中定义的别名
    3. 内置默认列名匹配

    Args:
        file_type: 文件类型
        header: CSV 表头列表
        user_aliases: 用户别名映射 {别名列名: 标准字段}
        extra_col_map: 额外传入的列映射 {别名列名: 标准字段}，用于 --col-map
    """
    defaults = DEFAULT_COLUMN_MAPS[file_type]
    mapping: Dict[str, str] = {}
    all_mapped_standard = set()

    if extra_col_map:
        for alias_col, std_field in extra_col_map.items():
            detected = _detect_column(header, [alias_col])
            if detected:
                mapping[std_field] = detected
                all_mapped_standard.add(std_field)

    if user_aliases:
        for alias_col, std_field in user_aliases.items():
            if std_field in mapping:
                continue
            detected = _detect_column(header, [alias_col])
            if detected:
                mapping[std_field] = detected
                all_mapped_standard.add(std_field)

    for field_name, candidates in defaults.items():
        if field_name in mapping:
            continue
        detected = _detect_column(header, candidates)
        if detected:
            mapping[field_name] = detected
            all_mapped_standard.add(field_name)

    return mapping


def _parse_amount(raw: str, source_file: str, source_row: int) -> Tuple[Optional[float], Optional[ParseError]]:
    """解析金额，处理千分位、正负号、货币符号等."""
    if raw is None:
        return None, ParseError(source_file, source_row, "invalid_amount", "金额为空")
    s = str(raw).strip()
    if not s:
        return None, ParseError(source_file, source_row, "invalid_amount", "金额为空")
    cleaned = s.replace(",", "").replace("，", "").replace("¥", "").replace("￥", "").replace(" ", "")
    try:
        val = float(cleaned)
        return val, None
    except ValueError:
        return None, ParseError(source_file, source_row, "invalid_amount", f"非法金额: {raw}")


def _parse_rows(
    file_path: str,
    file_type: FileType,
    header: List[str],
    rows: Iterator[Dict[str, Any]],
    user_aliases: Optional[Dict[str, str]] = None,
    extra_col_map: Optional[Dict[str, str]] = None,
) -> Tuple[ParseResult, ImportedFile]:
    """共享的行解析逻辑. CSV 和 XLSX 都走这条路径.

    Args:
        file_path: 源文件路径（用于溯源和错误信息）
        file_type: 文件类型
        header: 表头列名列表
        rows: 行迭代器，每行是 dict {列名: 值}
        user_aliases: 用户别名配置
        extra_col_map: 命令行传入的列映射

    Returns:
        (解析结果, 导入文件记录)
    """
    from .config import STANDARD_FIELDS

    result = ParseResult()
    imported_file = ImportedFile(
        file_type=file_type,
        file_path=os.path.abspath(file_path),
    )

    if extra_col_map:
        for alias_col, std_field in extra_col_map.items():
            if std_field not in STANDARD_FIELDS:
                raise ValueError(
                    f"列映射 '{alias_col}' -> '{std_field}' 无效。"
                    f"可用标准字段: {sorted(STANDARD_FIELDS)}。"
                    f"当前文件类型支持的列名: "
                    f"txn_id={DEFAULT_COLUMN_MAPS[file_type]['txn_id']}, "
                    f"amount={DEFAULT_COLUMN_MAPS[file_type]['amount']}, "
                    f"date={DEFAULT_COLUMN_MAPS[file_type]['date']}, "
                    f"counterparty={DEFAULT_COLUMN_MAPS[file_type]['counterparty']}, "
                    f"description={DEFAULT_COLUMN_MAPS[file_type]['description']}, "
                    f"currency={DEFAULT_COLUMN_MAPS[file_type]['currency']}"
                )

    col_map = _build_column_map(file_type, header, user_aliases, extra_col_map)

    if extra_col_map:
        for alias_col in extra_col_map:
            if alias_col not in header:
                all_header_cols = ", ".join(repr(h) for h in header)
                raise ValueError(
                    f"列映射 '{alias_col}' 在文件 {file_path} 的表头中未找到。"
                    f"现有表头列: {all_header_cols}"
                )

    if "txn_id" not in col_map:
        raise ValueError(
            f"文件 {file_path} 缺少交易号列. "
            f"支持的列名: {DEFAULT_COLUMN_MAPS[file_type]['txn_id']}"
            + (f"，已配置别名: {list(user_aliases.keys())}" if user_aliases else "")
            + (f"，--col-map: {list(extra_col_map.keys())}" if extra_col_map else "")
        )
    if "amount" not in col_map:
        raise ValueError(
            f"文件 {file_path} 缺少金额列. "
            f"支持的列名: {DEFAULT_COLUMN_MAPS[file_type]['amount']}"
            + (f"，已配置别名: {list(user_aliases.keys())}" if user_aliases else "")
            + (f"，--col-map: {list(extra_col_map.keys())}" if extra_col_map else "")
        )

    seen_ids: Dict[str, int] = {}

    for idx, row in enumerate(rows, start=2):
        raw_row = dict(row)
        raw_txn_id_raw = row.get(col_map["txn_id"])
        raw_txn_id = (str(raw_txn_id_raw).strip() if raw_txn_id_raw is not None else "")

        if not raw_txn_id:
            result.errors.append(ParseError(
                source_file=file_path,
                source_row=idx,
                error_type="missing_txn_id",
                message="缺少交易号",
                raw_row=raw_row,
            ))
            continue

        amount, err = _parse_amount(row.get(col_map["amount"], ""), file_path, idx)
        if err is not None:
            err.raw_row = raw_row
            result.errors.append(err)
            continue

        txn = Transaction(
            txn_id=raw_txn_id,
            amount=amount,
            date=(str(row.get(col_map["date"], "")).strip() if "date" in col_map else ""),
            file_type=file_type,
            source_file=file_path,
            source_row=idx,
            counterparty=(str(row.get(col_map["counterparty"], "")).strip() if "counterparty" in col_map else ""),
            description=(str(row.get(col_map["description"], "")).strip() if "description" in col_map else ""),
            currency=(str(row.get(col_map["currency"], "CNY")).strip() if "currency" in col_map else "CNY") or "CNY",
            raw_data=raw_row,
        )
        if raw_txn_id in seen_ids:
            result.duplicates.append(ParseError(
                source_file=file_path,
                source_row=idx,
                error_type="duplicate_txn_id",
                message=f"重复流水号 {raw_txn_id}, 首次出现于第 {seen_ids[raw_txn_id]} 行",
                raw_row=raw_row,
            ))
        else:
            seen_ids[raw_txn_id] = idx
        result.transactions.append(txn)

    imported_file.row_count = result.row_count
    imported_file.error_count = result.error_count
    return result, imported_file


def _load_user_aliases(
    storage_dir: Optional[str],
    file_type: FileType,
) -> Dict[str, str]:
    if storage_dir is None:
        return {}
    try:
        cfg = load_config(storage_dir)
        return get_column_aliases(cfg, file_type)
    except Exception:
        return {}


def parse_csv(
    file_path: str,
    file_type: FileType,
    encoding: str = "utf-8-sig",
    storage_dir: Optional[str] = None,
    extra_col_map: Optional[Dict[str, str]] = None,
) -> Tuple[ParseResult, ImportedFile]:
    """解析CSV文件.

    Args:
        file_path: CSV 文件路径
        file_type: 文件类型
        encoding: 文件编码
        storage_dir: 存储目录，用于读取 config.yaml 中的列名别名配置（可选）
        extra_col_map: 额外列映射 {别名列名: 标准字段}，来自 --col-map

    Returns:
        (解析结果, 导入文件记录)
    """
    if not os.path.isfile(file_path):
        raise FileNotFoundError(f"文件不存在: {file_path}")

    user_aliases = _load_user_aliases(storage_dir, file_type)

    with open(file_path, "r", encoding=encoding, newline="") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            raise ValueError(f"CSV 文件无表头: {file_path}")

        header = list(reader.fieldnames)
        return _parse_rows(file_path, file_type, header, reader, user_aliases, extra_col_map)


def parse_xlsx(
    file_path: str,
    file_type: FileType,
    storage_dir: Optional[str] = None,
    extra_col_map: Optional[Dict[str, str]] = None,
) -> Tuple[ParseResult, ImportedFile]:
    """解析 XLSX 文件（读取第一个 sheet，走与 CSV 完全相同的列名匹配+行解析逻辑）.

    Args:
        file_path: XLSX 文件路径
        file_type: 文件类型
        storage_dir: 存储目录，用于读取 config.yaml 中的列名别名配置（可选）
        extra_col_map: 额外列映射 {别名列名: 标准字段}，来自 --col-map

    Returns:
        (解析结果, 导入文件记录)
    """
    if not os.path.isfile(file_path):
        raise FileNotFoundError(f"文件不存在: {file_path}")

    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ValueError("未安装 openpyxl，无法解析 xlsx 文件。请 pip install openpyxl")

    user_aliases = _load_user_aliases(storage_dir, file_type)

    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        iter_rows = ws.iter_rows(values_only=True)

        try:
            header_row = next(iter_rows)
        except StopIteration:
            raise ValueError(f"XLSX 文件无表头: {file_path}")

        header = [("" if h is None else str(h)).strip() for h in header_row]
        if not any(header):
            raise ValueError(f"XLSX 文件无表头: {file_path}")

        def _row_gen() -> Iterator[Dict[str, Any]]:
            for values in iter_rows:
                if values is None:
                    continue
                if all(v is None or (isinstance(v, str) and not v.strip()) for v in values):
                    continue
                row_dict: Dict[str, Any] = {}
                for i, col_name in enumerate(header):
                    if i < len(values):
                        v = values[i]
                        if isinstance(v, str):
                            v = v.strip()
                        row_dict[col_name] = v
                    else:
                        row_dict[col_name] = ""
                yield row_dict

        return _parse_rows(file_path, file_type, header, _row_gen(), user_aliases, extra_col_map)
    finally:
        wb.close()


def parse_file(
    file_path: str,
    file_type: FileType,
    encoding: str = "utf-8-sig",
    storage_dir: Optional[str] = None,
    extra_col_map: Optional[Dict[str, str]] = None,
) -> Tuple[ParseResult, ImportedFile]:
    """统一入口：根据扩展名自动选择 CSV 或 XLSX 解析器."""
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".xlsx":
        return parse_xlsx(file_path, file_type, storage_dir=storage_dir, extra_col_map=extra_col_map)
    elif ext == ".csv":
        return parse_csv(file_path, file_type, encoding=encoding, storage_dir=storage_dir, extra_col_map=extra_col_map)
    else:
        raise ValueError(
            f"不支持的文件扩展名: {ext}。仅支持 .csv 和 .xlsx"
        )
